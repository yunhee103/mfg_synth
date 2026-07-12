"""시스템별 출력.

ERP: MariaDB용 SQL 덤프 (스키마 + INSERT).
MES: 월별 CSV (하이픈 없는 코드, 결측, 중복 포함).
품질검사: 월별 Excel (병합 제목, 3행 헤더, 날짜 포맷 혼재).
ground truth: 심어둔 이상과 오염의 명세 JSON.
"""

import json
from datetime import date
from pathlib import Path
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .config import LEGACY_CODE_MAP, SimConfig
from .corrupt import inspector_date_style, messy_date, typo_name
from .master import MasterData
from .simulate import SimResult

_ERP_SCHEMA = """
SET NAMES utf8mb4;
CREATE DATABASE IF NOT EXISTS erp_source DEFAULT CHARACTER SET utf8mb4;
USE erp_source;

DROP TABLE IF EXISTS customers, suppliers, items, bom,
    sales_order_h, sales_order_d, purchase_orders,
    material_receipts, shipments, inventory_daily, item_change_log;

CREATE TABLE customers (
    customer_code VARCHAR(10) PRIMARY KEY,
    customer_name VARCHAR(50),
    region VARCHAR(20)
);
CREATE TABLE suppliers (
    supplier_code VARCHAR(10) PRIMARY KEY,
    supplier_name VARCHAR(50)
);
CREATE TABLE items (
    item_code VARCHAR(20) PRIMARY KEY,
    item_name VARCHAR(100),
    item_type VARCHAR(10),
    item_group VARCHAR(30),
    unit_price INT,
    supplier_code VARCHAR(10),
    lead_time_days INT
);
CREATE TABLE bom (
    parent_item VARCHAR(20),
    child_item VARCHAR(20),
    qty_per INT,
    PRIMARY KEY (parent_item, child_item)
);
CREATE TABLE sales_order_h (
    order_no VARCHAR(20) PRIMARY KEY,
    customer_code VARCHAR(10),
    order_date DATE
);
CREATE TABLE sales_order_d (
    order_no VARCHAR(20),
    line_no INT,
    item_code VARCHAR(20),
    qty INT,
    due_date DATE,
    PRIMARY KEY (order_no, line_no)
);
CREATE TABLE purchase_orders (
    po_no VARCHAR(20) PRIMARY KEY,
    supplier_code VARCHAR(10),
    material_code VARCHAR(20),
    qty INT,
    order_date DATE,
    promised_date DATE
);
CREATE TABLE material_receipts (
    po_no VARCHAR(20),
    material_code VARCHAR(20),
    receipt_date DATE,
    qty INT
);
CREATE TABLE shipments (
    ship_no VARCHAR(20) PRIMARY KEY,
    order_no VARCHAR(20),
    line_no INT,
    item_code VARCHAR(20),
    qty_shipped INT,
    ship_date DATE
);
CREATE TABLE inventory_daily (
    snap_date DATE,
    material_code VARCHAR(20),
    qty_on_hand INT,
    qty_on_order INT,
    PRIMARY KEY (snap_date, material_code)
);
CREATE TABLE item_change_log (
    log_id INT PRIMARY KEY AUTO_INCREMENT,
    change_date DATE,
    item_code VARCHAR(20),
    field_name VARCHAR(30),
    old_value VARCHAR(100),
    new_value VARCHAR(100),
    changed_by VARCHAR(30)
);
"""


def _sql_val(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, (int, float, np.integer, np.floating)):
        return str(v)
    if isinstance(v, date):
        return f"'{v:%Y-%m-%d}'"
    return "'" + str(v).replace("'", "''") + "'"


def _inserts(table: str, rows: List[dict], batch: int = 500) -> List[str]:
    if not rows:
        return []
    cols = list(rows[0].keys())
    lines = []
    for i in range(0, len(rows), batch):
        chunk = rows[i:i + batch]
        values = ",\n".join(
            "(" + ", ".join(_sql_val(r[c]) for c in cols) + ")"
            for r in chunk)
        lines.append(
            f"INSERT INTO {table} ({', '.join(cols)}) VALUES\n{values};\n")
    return lines


def write_erp_sql(path: Path, md: MasterData, res: SimResult,
                  erp_shipments: List[dict], cfg: SimConfig) -> None:
    """ERP 소스 DB용 SQL 덤프를 생성한다."""
    items = [
        {"item_code": f.item_code, "item_name": f.item_name,
         "item_type": "FG", "item_group": f.item_group,
         "unit_price": f.unit_price, "supplier_code": None,
         "lead_time_days": None}
        for f in md.finished_goods.values()
    ] + [
        {"item_code": m.item_code, "item_name": m.item_name,
         "item_type": "RM", "item_group": "자재",
         "unit_price": m.unit_cost, "supplier_code": m.supplier_code,
         "lead_time_days": m.lead_time_days}
        for m in md.materials.values()
    ]
    bom_rows = [
        {"parent_item": p, "child_item": c, "qty_per": q}
        for p, comps in md.bom.items() for c, q in comps.items()
    ]
    parts = [_ERP_SCHEMA]
    parts += _inserts("customers", md.customers)
    parts += _inserts("suppliers", md.suppliers)
    parts += _inserts("items", items)
    parts += _inserts("bom", bom_rows)
    parts += _inserts("sales_order_h", res.sales_order_h)
    parts += _inserts("sales_order_d", res.sales_order_d)
    parts += _inserts("purchase_orders", res.purchase_orders)
    parts += _inserts("material_receipts", res.material_receipts)
    parts += _inserts("shipments", erp_shipments)
    parts += _inserts("inventory_daily", res.inventory_daily)
    change_log = [
        {"log_id": i + 1, "change_date": cfg.reorg.change_date, "item_code": f.item_code,
         "field_name": "item_group", "old_value": f.item_group_old,
         "new_value": f.item_group, "changed_by": "admin"}
        for i, f in enumerate(
            f for f in md.finished_goods.values() if f.item_group_old)
    ]
    parts += _inserts("item_change_log", change_log)
    path.write_text("\n".join(parts), encoding="utf-8")


def write_mes_csv(out_dir: Path, wo_rows: List[dict],
                  prod_rows: List[dict]) -> None:
    """MES 데이터를 월별 CSV로 출력한다. 컬럼명은 MES 벤더 스타일."""
    out_dir.mkdir(parents=True, exist_ok=True)

    wo_df = pd.DataFrame(wo_rows).rename(columns={
        "wo_no": "WORK_ORD_NO", "order_no": "REF_ORD_NO",
        "item_code": "ITEM_CD", "equip_code": "EQP_CD",
        "qty": "ORD_QTY", "create_date": "REG_DT",
        "release_date": "REL_DT", "complete_date": "CMPL_DT",
    })
    wo_df.to_csv(out_dir / "mes_work_orders.csv", index=False,
                 encoding="utf-8-sig")

    prod_df = pd.DataFrame(prod_rows).rename(columns={
        "wo_no": "WORK_ORD_NO", "work_date": "WORK_DT",
        "equip_code": "EQP_CD", "item_code": "ITEM_CD",
        "qty_produced": "PROD_QTY", "qty_defect": "DEF_QTY",
        "cycle_time_s": "CYCLE_TM", "worker": "WORKER_NM",
    })
    prod_df["WORK_DT"] = pd.to_datetime(prod_df["WORK_DT"])
    for ym, grp in prod_df.groupby(prod_df["WORK_DT"].dt.strftime("%Y%m")):
        grp.to_csv(out_dir / f"mes_prod_result_{ym}.csv", index=False,
                   encoding="utf-8-sig")


def write_inspection_excel(out_dir: Path, insp_rows: List[dict],
                           cfg: SimConfig, rng: np.random.Generator) -> None:
    """품질검사 대장을 월별 Excel로 출력한다.

    수기 대장 이관을 흉내 낸다: 병합 제목 행, 3행 헤더,
    검사자별로 다른 날짜 표기, 이름 오타, 문자열 숫자.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(insp_rows)
    df["ym"] = pd.to_datetime(df["insp_date"]).dt.strftime("%Y%m")

    for ym, grp in df.groupby("ym"):
        wb = Workbook()
        ws = wb.active
        ws.title = "검사대장"
        ws.merge_cells("A1:H1")
        ws["A1"] = f"품질검사 대장 ({ym[:4]}년 {int(ym[4:])}월)"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["검사일", "작업지시번호", "품목코드", "LOT수량",
                   "샘플수량", "불량수", "판정", "검사자"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=h).font = Font(bold=True)

        row_idx = 4
        for _, r in grp.iterrows():
            style = inspector_date_style(r["inspector"])
            name = r["inspector"]
            if rng.random() < cfg.corruption.excel_inspector_typo_rate:
                name = typo_name(name, rng)
            lot_qty = (f"{r['lot_qty']:,}"
                       if rng.random() < cfg.corruption.excel_qty_as_text_rate
                       else int(r["lot_qty"]))
            values = [
                messy_date(r["insp_date"], style) if
                cfg.corruption.excel_date_chaos else str(r["insp_date"]),
                r["wo_no"], r["item_code"], lot_qty,
                int(r["sample_qty"]), int(r["defect_qty"]),
                r["judgment"], name,
            ]
            for col, v in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col, value=v)
            row_idx += 1
        wb.save(out_dir / f"quality_insp_{ym}.xlsx")


def write_ground_truth(path: Path, cfg: SimConfig, md: MasterData) -> None:
    """심어둔 이상과 오염의 명세를 JSON으로 남긴다.

    분석 결과 검증용이며, 분석 단계에서는 참조하지 않는다.
    """
    deg = cfg.degradation
    delay = cfg.supplier_delay
    affected_fg_equip = sorted(
        f.item_code for f in md.finished_goods.values()
        if f.equip_code == deg.equip_code)
    delayed_materials = sorted(
        m.item_code for m in md.materials.values()
        if m.supplier_code == delay.supplier_code)
    affected_fg_material = sorted({
        fg for fg, comps in md.bom.items()
        if any(m in comps for m in delayed_materials)})

    truth = {
        "seed": cfg.seed,
        "anomalies": {
            "chain_A_equipment_degradation": {
                "equip_code": deg.equip_code,
                "start": str(deg.start),
                "ramp_days": deg.ramp_days,
                "max_cycle_time_drift": deg.max_drift,
                "defect_uplift": deg.defect_uplift,
                "affected_finished_goods": affected_fg_equip,
            },
            "chain_B_supplier_delay": {
                "supplier_code": delay.supplier_code,
                "start": str(delay.start),
                "lead_time_change": [delay.old_lead_time,
                                     delay.new_lead_time],
                "delayed_materials": delayed_materials,
                "affected_finished_goods": affected_fg_material,
            },
            "overlap_items": sorted(
                set(affected_fg_equip) & set(affected_fg_material)),
        },
        "master_change": {
            "type": "item_group_reorg",
            "change_date": str(cfg.reorg.change_date),
            "new_group": cfg.reorg.new_group,
            "items": {
                f.item_code: f.item_group_old
                for f in md.finished_goods.values() if f.item_group_old
            },
        },
        "corruption": {
            "mes_strip_hyphen": cfg.corruption.mes_strip_hyphen,
            "mes_legacy_codes": LEGACY_CODE_MAP,
            "mes_missing_rate": cfg.corruption.mes_missing_rate,
            "mes_duplicate_rate": cfg.corruption.mes_duplicate_rate,
            "erp_ship_qty_typo_rate": cfg.corruption.erp_ship_qty_typo_rate,
            "excel": {
                "date_chaos": cfg.corruption.excel_date_chaos,
                "inspector_typo_rate":
                    cfg.corruption.excel_inspector_typo_rate,
                "qty_as_text_rate": cfg.corruption.excel_qty_as_text_rate,
            },
        },
    }
    path.write_text(json.dumps(truth, ensure_ascii=False, indent=2),
                    encoding="utf-8")
